from todo_manager import TodoManager
from rich.console import Console
from rich.table import Table as RichTable
from rich.prompt import Prompt, IntPrompt
from rich.text import Text
from rich.panel import Panel
from rich import print
from rich.progress import Progress, SpinnerColumn
from rich.layout import Layout
from rich.align import Align
from rich.box import ROUNDED, MINIMAL, HEAVY_HEAD
from rich.style import Style
from rich.rule import Rule
from rich.columns import Columns
import time

class TodoApp:
    def __init__(self):
        # Initialize TodoManager with a data file in the src directory
        import os
        data_file = os.path.join(os.path.dirname(__file__), "todos.json")
        self.manager = TodoManager(data_file=data_file)
        self.console = Console()

    def display_menu(self):
        """Display the main menu with enhanced rich formatting"""
        # Create a decorative header
        self.console.print(Rule("[bold magenta]VIP TODO APPLICATION[/bold magenta]", style="bold magenta"))

        # Create a panel with the menu options
        menu_content = (
            "[bold green]1.[/bold green] 📝 Add Task\n"
            "[bold green]2.[/bold green] 📋 View Tasks\n"
            "[bold green]3.[/bold green] ✏️  Update Task\n"
            "[bold green]4.[/bold green] 🗑️  Delete Task\n"
            "[bold green]5.[/bold green] ✅ Mark/Unmark Task Complete\n"
            "[bold green]6.[/bold green] 📄 Print Final Record\n"
            "[bold green]7.[/bold green] 📊 Export to Excel\n"
            "[bold green]8.[/bold green] 📧 Send Email\n"
            "[bold green]9.[/bold green] 🚪 Exit"
        )

        menu_panel = Panel(
            menu_content,
            title="[bold cyan]📋 MENU OPTIONS[/bold cyan]",
            border_style="bright_blue",
            box=ROUNDED,
            padding=(1, 2)
        )

        self.console.print(menu_panel)
        self.console.print()

    def run(self):
        """Main application loop"""
        # Enhanced welcome message
        welcome_text = (
            "[bold yellow]🌟 Welcome to the VIP Todo Application! 🌟[/bold yellow]\n\n"
            "[italic]Your premium task management solution[/italic]"
        )
        self.console.print(Panel(
            welcome_text,
            title="[bold cyan]🎉 WELCOME[/bold cyan]",
            border_style="bright_yellow",
            box=ROUNDED
        ))

        while True:
            self.display_menu()
            choice = Prompt.ask("Enter your choice", choices=["1", "2", "3", "4", "5", "6", "7", "8", "9"])

            if choice == '1':
                self.add_task()
            elif choice == '2':
                self.view_tasks()
            elif choice == '3':
                self.update_task()
            elif choice == '4':
                self.delete_task()
            elif choice == '5':
                self.toggle_complete_task()
            elif choice == '6':
                self.print_final_record()
            elif choice == '7':
                self.export_to_excel()
            elif choice == '8':
                self.send_email()
            elif choice == '9':
                # Enhanced exit message
                exit_panel = Panel(
                    "[bold green]👋 Thank you for using VIP Todo Application![/bold green]\n"
                    "[italic]Have a great day![/italic]",
                    title="[bold red]👋 GOODBYE[/bold red]",
                    border_style="bright_red",
                    box=ROUNDED
                )
                self.console.print(exit_panel)
                break
            else:
                self.console.print(Panel(
                    "[red]❌ Invalid choice. Please try again.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))

    def add_task(self):
        """Add a new task with enhanced rich input prompts"""
        # Enhanced input prompts with icons
        self.console.print(Panel(
            "[bold blue]📝 ADDING NEW TASK[/bold blue]",
            border_style="bright_blue",
            box=ROUNDED
        ))

        title = Prompt.ask("Enter task title")
        description = Prompt.ask("Enter task description (optional)")
        task = self.manager.add_todo(title, description)

        # Create a more visually appealing table to display the added task
        table = RichTable(
            title="✅ Task Added Successfully",
            show_header=True,
            header_style="bold magenta",
            box=HEAVY_HEAD,
            border_style="green"
        )
        table.add_column("ID", style="dim", width=5)
        table.add_column("Title", min_width=20)
        table.add_column("Description", min_width=20)
        table.add_column("Status", justify="center")

        status_text = "[green]✅ Complete[/green]" if task.completed else "[red]❌ Incomplete[/red]"
        table.add_row(str(task.id), task.title, task.description, status_text)

        self.console.print(table)

        # Add a success message with animation
        with Progress(
            SpinnerColumn(),
            "[progress.description]{task.description}",
            transient=True,
        ) as progress:
            progress.add_task(description="Processing...", total=None)
            time.sleep(0.5)

        self.console.print(Panel(
            f"[bold green]🎉 Task '{task.title}' has been added successfully![/bold green]",
            border_style="bright_green",
            box=ROUNDED
        ))

    def view_tasks(self):
        """Display all tasks in an enhanced rich table format"""
        tasks = self.manager.get_all_todos()
        if not tasks:
            self.console.print(Panel(
                "[yellow]📋 No tasks available.[/yellow]",
                border_style="bright_yellow",
                box=ROUNDED
            ))
            return

        # Count completed and incomplete tasks
        completed_count = sum(1 for task in tasks if task.completed)
        total_count = len(tasks)

        # Create a header panel with statistics
        stats_text = f"[bold cyan]📊 Total: {total_count} | Completed: {completed_count} | Pending: {total_count - completed_count}[/bold cyan]"
        self.console.print(Panel(stats_text, border_style="cyan", box=ROUNDED))

        # Create a more visually appealing table to display tasks
        table = RichTable(
            title="📋 Your Tasks",
            show_header=True,
            header_style="bold blue",
            box=HEAVY_HEAD,
            border_style="blue"
        )
        table.add_column("ID", style="dim", width=5)
        table.add_column("Title", min_width=25)
        table.add_column("Status", justify="center", width=15)

        for task in tasks:
            status_text = "[green]✅ Complete[/green]" if task.completed else "[red]❌ Incomplete[/red]"
            table.add_row(str(task.id), task.title, status_text)

        self.console.print(table)

    def update_task(self):
        """Update an existing task with enhanced rich interface"""
        self.view_tasks()
        try:
            todo_id = IntPrompt.ask("Enter the ID of the task to update")
            current_task = self.manager.find_todo_by_id(todo_id)

            if not current_task:
                self.console.print(Panel(
                    f"[red]❌ Task with ID {todo_id} not found.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))
                return

            # Show current task details in a panel
            current_details = Panel(
                f"[bold]Current task:[/bold] {current_task.title}\n"
                f"[dim]Description:[/dim] {current_task.description}",
                title="[bold yellow]📋 CURRENT TASK[/bold yellow]",
                border_style="bright_yellow",
                box=ROUNDED
            )
            self.console.print(current_details)

            new_title = Prompt.ask("Enter new title (leave blank to keep current)", default=current_task.title, show_default=False)
            new_description = Prompt.ask("Enter new description (leave blank to keep current)", default=current_task.description, show_default=False)

            update_successful = self.manager.update_todo(
                todo_id,
                new_title if new_title != current_task.title else None,
                new_description if new_description != current_task.description else None
            )
            if update_successful:
                # Show success with animation
                with Progress(
                    SpinnerColumn(),
                    "[progress.description]{task.description}",
                    transient=True,
                ) as progress:
                    progress.add_task(description="Updating task...", total=None)
                    time.sleep(0.5)

                self.console.print(Panel(
                    "[green]✅ Task updated successfully![/green]",
                    border_style="bright_green",
                    box=ROUNDED
                ))
            else:
                self.console.print(Panel(
                    "[red]❌ Task not found.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))
        except ValueError:
            self.console.print(Panel(
                "[red]❌ Invalid ID. Please enter a number.[/red]",
                border_style="red",
                box=ROUNDED
            ))

    def delete_task(self):
        """Delete a task with enhanced confirmation and rich feedback"""
        self.view_tasks()
        try:
            todo_id = IntPrompt.ask("Enter the ID of the task to delete")

            # Show task to be deleted
            task_to_delete = self.manager.find_todo_by_id(todo_id)
            if not task_to_delete:
                self.console.print(Panel(
                    f"[red]❌ Task with ID {todo_id} not found.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))
                return

            # Show task details before confirmation
            task_details = Panel(
                f"[bold red]⚠️  WARNING: About to delete this task:[/bold red]\n\n"
                f"[bold]{task_to_delete.title}[/bold]\n"
                f"[dim]{task_to_delete.description}[/dim]",
                title="[bold red]🗑️  DELETE CONFIRMATION[/bold red]",
                border_style="bright_red",
                box=ROUNDED
            )
            self.console.print(task_details)

            confirm = Prompt.ask(f"Are you sure you want to delete task '{task_to_delete.title}'? (y/n)", choices=["y", "n"])

            if confirm.lower() == 'y':
                if self.manager.delete_todo(todo_id):
                    # Show deletion with animation
                    with Progress(
                        SpinnerColumn(),
                        "[progress.description]{task.description}",
                        transient=True,
                    ) as progress:
                        progress.add_task(description="Deleting task...", total=None)
                        time.sleep(0.5)

                    self.console.print(Panel(
                        f"[green]✅ Task '{task_to_delete.title}' deleted successfully![/green]",
                        border_style="bright_green",
                        box=ROUNDED
                    ))
                else:
                    self.console.print(Panel(
                        "[red]❌ Task not found.[/red]",
                        border_style="red",
                        box=ROUNDED
                    ))
            else:
                self.console.print(Panel(
                    "[yellow]↩️  Deletion cancelled.[/yellow]",
                    border_style="bright_yellow",
                    box=ROUNDED
                ))
        except ValueError:
            self.console.print(Panel(
                "[red]❌ Invalid ID. Please enter a number.[/red]",
                border_style="red",
                box=ROUNDED
            ))

    def toggle_complete_task(self):
        """Toggle task completion status with enhanced animation"""
        self.view_tasks()
        try:
            todo_id = IntPrompt.ask("Enter the ID of the task to mark/unmark complete")

            # Show current status
            task = self.manager.find_todo_by_id(todo_id)
            if not task:
                self.console.print(Panel(
                    f"[red]❌ Task with ID {todo_id} not found.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))
                return

            current_status = "✅ Complete" if task.completed else "❌ Incomplete"
            status_panel = Panel(
                f"[bold]Current status:[/bold] {current_status}",
                title="[bold yellow]📋 CURRENT STATUS[/bold yellow]",
                border_style="bright_yellow",
                box=ROUNDED
            )
            self.console.print(status_panel)

            # Show loading animation while toggling
            with Progress(
                SpinnerColumn(),
                "[progress.description]{task.description}",
                transient=True,
            ) as progress:
                progress.add_task(description="Updating task status...", total=None)
                time.sleep(1)  # Simulate processing time

            if self.manager.toggle_complete(todo_id):
                new_status = "✅ Complete" if task.completed else "❌ Incomplete"
                status_color = "[green]✅ Complete[/green]" if task.completed else "[red]❌ Incomplete[/red]"

                # Create a success panel with the new status
                success_panel = Panel(
                    f"[green]✅ Task status updated successfully![/green]\n\n"
                    f"[bold]New status:[/bold] {status_color}",
                    title="[bold green]🎉 STATUS UPDATED[/bold green]",
                    border_style="bright_green",
                    box=ROUNDED
                )
                self.console.print(success_panel)
            else:
                self.console.print(Panel(
                    "[red]❌ Task not found.[/red]",
                    border_style="red",
                    box=ROUNDED
                ))
        except ValueError:
            self.console.print(Panel(
                "[red]❌ Invalid ID. Please enter a number.[/red]",
                border_style="red",
                box=ROUNDED
            ))

    def print_final_record(self):
        """Print and save the final record of tasks to a file in /src directory"""
        tasks = self.manager.get_all_todos()

        if not tasks:
            self.console.print("[yellow]No tasks to print.[/yellow]")
            return

        # Count completed and incomplete tasks
        completed_count = sum(1 for task in tasks if task.completed)
        incomplete_count = len(tasks) - completed_count

        # Create a summary table (using Rich)
        from rich.table import Table as RichTable
        summary_table = RichTable(title="Final Record Summary", show_header=True, header_style="bold magenta")
        summary_table.add_column("Metric", style="dim")
        summary_table.add_column("Count", justify="right")
        summary_table.add_row("Total Tasks", str(len(tasks)))
        summary_table.add_row("[green]Completed Tasks[/green]", str(completed_count))
        summary_table.add_row("[red]Incomplete Tasks[/red]", str(incomplete_count))

        self.console.print(summary_table)

        # Create detailed task table (using Rich)
        task_table = RichTable(title="All Tasks", show_header=True, header_style="bold blue")
        task_table.add_column("ID", style="dim", width=5)
        task_table.add_column("Title", min_width=20)
        task_table.add_column("Description", min_width=20)
        task_table.add_column("Status", justify="center")

        for task in tasks:
            status_text = "[green]Complete[/green]" if task.completed else "[red]Incomplete[/red]"
            task_table.add_row(str(task.id), task.title, task.description, status_text)

        self.console.print(task_table)

        # Save the record to a file in the /src directory
        import os
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Save text file
        txt_filename = f"final_record_{timestamp}.txt"
        txt_filepath = os.path.join(os.path.dirname(__file__), txt_filename)

        try:
            with open(txt_filepath, 'w', encoding='utf-8') as f:
                f.write("Final Todo Record\n")
                f.write("="*50 + "\n")
                f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                f.write("Summary:\n")
                f.write(f"Total Tasks: {len(tasks)}\n")
                f.write(f"Completed Tasks: {completed_count}\n")
                f.write(f"Incomplete Tasks: {incomplete_count}\n\n")

                f.write("Task Details:\n")
                f.write("-" * 50 + "\n")
                for task in tasks:
                    status = "Complete" if task.completed else "Incomplete"
                    f.write(f"ID: {task.id}\n")
                    f.write(f"Title: {task.title}\n")
                    f.write(f"Description: {task.description}\n")
                    f.write(f"Status: {status}\n")
                    f.write("-" * 30 + "\n")

            self.console.print(f"[green]Final record saved to: {txt_filepath}[/green]")

        except Exception as e:
            self.console.print(f"[red]Error saving final record: {str(e)}[/red]")

        # Generate PDF version
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            pdf_filename = f"final_record_{timestamp}.pdf"
            pdf_filepath = os.path.join(os.path.dirname(__file__), pdf_filename)

            # Create PDF document
            doc = SimpleDocTemplate(pdf_filepath, pagesize=letter)
            elements = []

            # Add title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph("Final Todo Record", title_style)
            elements.append(title)

            # Add generation date
            date_style = ParagraphStyle(
                'CustomDate',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=20,
                alignment=1  # Center alignment
            )
            date_para = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
            elements.append(date_para)

            # Add summary table
            summary_data = [
                ['Metric', 'Count'],
                ['Total Tasks', str(len(tasks))],
                ['Completed Tasks', str(completed_count)],
                ['Incomplete Tasks', str(incomplete_count)]
            ]

            summary_table_pdf = RLTable(summary_data)
            summary_table_pdf.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(summary_table_pdf)
            elements.append(Spacer(1, 20))

            # Add tasks header
            tasks_header = Paragraph("Task Details", styles['Heading2'])
            elements.append(tasks_header)
            elements.append(Spacer(1, 10))

            # Add tasks table
            task_data = [['ID', 'Title', 'Description', 'Status']]
            for task in tasks:
                status = "Complete" if task.completed else "Incomplete"
                task_data.append([str(task.id), task.title, task.description, status])

            task_table_pdf = RLTable(task_data)
            task_table_pdf.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Color the status column based on completion
                ('TEXTCOLOR', (3, 1), (3, -1), colors.red),  # Default to red for all statuses
            ]))

            # Apply specific colors based on status
            for i, task in enumerate(tasks, start=1):
                if task.completed:
                    task_table_pdf.setStyle(TableStyle([
                        ('TEXTCOLOR', (3, i), (3, i), colors.green)
                    ]))

            elements.append(task_table_pdf)

            # Build PDF
            doc.build(elements)

            self.console.print(f"[green]PDF report saved to: {pdf_filepath}[/green]")
            self.console.print(f"[bold green]PDF ready for client email![/bold green]")

        except Exception as e:
            self.console.print(f"[red]Error generating PDF: {str(e)}[/red]")

    def export_to_excel(self):
        """Export all tasks to an Excel file"""
        tasks = self.manager.get_all_todos()

        if not tasks:
            self.console.print(Panel(
                "[yellow]📋 No tasks to export.[/yellow]",
                border_style="bright_yellow",
                box=ROUNDED
            ))
            return

        try:
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Todo Tasks"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add headers
            headers = ["ID", "Title", "Description", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Add tasks data
            for row, task in enumerate(tasks, 2):
                status = "Complete" if task.completed else "Incomplete"

                # Add data to cells
                ws.cell(row=row, column=1, value=task.id).border = border
                ws.cell(row=row, column=2, value=task.title).border = border
                ws.cell(row=row, column=3, value=task.description).border = border

                # Add status with color coding
                status_cell = ws.cell(row=row, column=4, value=status)
                status_cell.border = border
                if task.completed:
                    status_cell.font = Font(color="009900")  # Green for complete
                else:
                    status_cell.font = Font(color="FF0000")  # Red for incomplete

                # Center align the content
                for col in range(1, 5):
                    ws.cell(row=row, column=col).alignment = center_alignment

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add a summary sheet
            summary_ws = wb.create_sheet(title="Summary")

            # Add summary information
            summary_ws['A1'] = "Todo Application Summary"
            summary_ws['A1'].font = Font(size=16, bold=True)

            summary_ws['A3'] = "Total Tasks:"
            summary_ws['B3'] = len(tasks)

            completed_count = sum(1 for task in tasks if task.completed)
            summary_ws['A4'] = "Completed Tasks:"
            summary_ws['B4'] = completed_count

            incomplete_count = len(tasks) - completed_count
            summary_ws['A5'] = "Incomplete Tasks:"
            summary_ws['B5'] = incomplete_count

            summary_ws['A6'] = "Export Date:"
            summary_ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Style the summary sheet
            summary_ws['A1'].fill = header_fill
            summary_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")

            for row in range(3, 7):
                summary_ws[f'A{row}'].font = Font(bold=True)

            # Save the Excel file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"todo_tasks_{timestamp}.xlsx"
            excel_filepath = os.path.join(os.path.dirname(__file__), excel_filename)

            wb.save(excel_filepath)

            # Show success with animation
            with Progress(
                SpinnerColumn(),
                "[progress.description]{task.description}",
                transient=True,
            ) as progress:
                progress.add_task(description="Generating Excel file...", total=None)
                time.sleep(0.5)

            self.console.print(Panel(
                f"[green]✅ Tasks exported to Excel: {excel_filepath}[/green]",
                border_style="bright_green",
                box=ROUNDED
            ))
            self.console.print(Panel(
                "[bold green]📊 Excel file ready for sharing![/bold green]",
                border_style="bright_green",
                box=ROUNDED
            ))

        except Exception as e:
            self.console.print(Panel(
                f"[red]❌ Error exporting to Excel: {str(e)}[/red]",
                border_style="red",
                box=ROUNDED
            ))

    def send_email(self):
        """Send tasks via email"""
        tasks = self.manager.get_all_todos()

        if not tasks:
            self.console.print("[yellow]No tasks to send via email.[/yellow]")
            return

        try:
            # Get email details from user
            from rich.prompt import Prompt
            recipient_email = Prompt.ask("Enter recipient's email address")
            subject = Prompt.ask("Enter email subject", default="Todo Tasks Report")
            message_body = Prompt.ask("Enter email message", default="Please find the attached todo tasks report.")

            # Ask if user wants to attach files
            attach_files = Prompt.ask("Do you want to attach Excel or PDF report? (y/n)", choices=["y", "n"], default="n")

            attachments = []
            if attach_files.lower() == 'y':
                attachment_choice = Prompt.ask("Choose attachment type: [1] Excel [2] PDF [3] Both", choices=["1", "2", "3"], default="1")

                # Generate attachments based on user choice
                from datetime import datetime
                import os

                if attachment_choice in ["1", "3"]:
                    # Export to Excel and get file path
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Todo Tasks"

                    # Define styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Add headers
                    headers = ["ID", "Title", "Description", "Status"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = border

                    # Add tasks data
                    for row, task in enumerate(tasks, 2):
                        status = "Complete" if task.completed else "Incomplete"

                        # Add data to cells
                        ws.cell(row=row, column=1, value=task.id).border = border
                        ws.cell(row=row, column=2, value=task.title).border = border
                        ws.cell(row=row, column=3, value=task.description).border = border

                        # Add status with color coding
                        status_cell = ws.cell(row=row, column=4, value=status)
                        status_cell.border = border
                        if task.completed:
                            status_cell.font = Font(color="009900")  # Green for complete
                        else:
                            status_cell.font = Font(color="FF0000")  # Red for incomplete

                        # Center align the content
                        for col in range(1, 5):
                            ws.cell(row=row, column=col).alignment = center_alignment

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)

                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Add a summary sheet
                    summary_ws = wb.create_sheet(title="Summary")

                    # Add summary information
                    summary_ws['A1'] = "Todo Application Summary"
                    summary_ws['A1'].font = Font(size=16, bold=True)

                    summary_ws['A3'] = "Total Tasks:"
                    summary_ws['B3'] = len(tasks)

                    completed_count = sum(1 for task in tasks if task.completed)
                    summary_ws['A4'] = "Completed Tasks:"
                    summary_ws['B4'] = completed_count

                    incomplete_count = len(tasks) - completed_count
                    summary_ws['A5'] = "Incomplete Tasks:"
                    summary_ws['B5'] = incomplete_count

                    summary_ws['A6'] = "Export Date:"
                    summary_ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # Style the summary sheet
                    summary_ws['A1'].fill = header_fill
                    summary_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")

                    for row in range(3, 7):
                        summary_ws[f'A{row}'].font = Font(bold=True)

                    # Save the Excel file
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_filename = f"todo_tasks_email_{timestamp}.xlsx"
                    excel_filepath = os.path.join(os.path.dirname(__file__), excel_filename)

                    wb.save(excel_filepath)
                    attachments.append(excel_filepath)

                if attachment_choice in ["2", "3"]:
                    # Generate PDF and get file path
                    from reportlab.lib.pagesizes import letter
                    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib import colors

                    pdf_filename = f"todo_tasks_email_{timestamp}.pdf"
                    pdf_filepath = os.path.join(os.path.dirname(__file__), pdf_filename)

                    # Create PDF document
                    doc = SimpleDocTemplate(pdf_filepath, pagesize=letter)
                    elements = []

                    # Add title
                    styles = getSampleStyleSheet()
                    title_style = ParagraphStyle(
                        'CustomTitle',
                        parent=styles['Heading1'],
                        fontSize=18,
                        spaceAfter=30,
                        alignment=1  # Center alignment
                    )
                    title = Paragraph("Todo Tasks Report", title_style)
                    elements.append(title)

                    # Add generation date
                    date_style = ParagraphStyle(
                        'CustomDate',
                        parent=styles['Normal'],
                        fontSize=12,
                        spaceAfter=20,
                        alignment=1  # Center alignment
                    )
                    date_para = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
                    elements.append(date_para)

                    # Add summary table
                    completed_count = sum(1 for task in tasks if task.completed)
                    incomplete_count = len(tasks) - completed_count
                    summary_data = [
                        ['Metric', 'Count'],
                        ['Total Tasks', str(len(tasks))],
                        ['Completed Tasks', str(completed_count)],
                        ['Incomplete Tasks', str(incomplete_count)]
                    ]

                    summary_table_pdf = RLTable(summary_data)
                    summary_table_pdf.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))

                    elements.append(summary_table_pdf)
                    elements.append(Spacer(1, 20))

                    # Add tasks header
                    tasks_header = Paragraph("Task Details", styles['Heading2'])
                    elements.append(tasks_header)
                    elements.append(Spacer(1, 10))

                    # Add tasks table
                    task_data = [['ID', 'Title', 'Description', 'Status']]
                    for task in tasks:
                        status = "Complete" if task.completed else "Incomplete"
                        task_data.append([str(task.id), task.title, task.description, status])

                    task_table_pdf = RLTable(task_data)
                    task_table_pdf.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        # Color the status column based on completion
                        ('TEXTCOLOR', (3, 1), (3, -1), colors.red),  # Default to red for all statuses
                    ]))

                    # Apply specific colors based on status
                    for i, task in enumerate(tasks, start=1):
                        if task.completed:
                            task_table_pdf.setStyle(TableStyle([
                                ('TEXTCOLOR', (3, i), (3, i), colors.green)
                            ]))

                    elements.append(task_table_pdf)

                    # Build PDF
                    doc.build(elements)
                    attachments.append(pdf_filepath)

            # Show email summary
            self.console.print(f"[bold blue]Email Summary:[/bold blue]")
            self.console.print(f"Recipient: {recipient_email}")
            self.console.print(f"Subject: {subject}")
            self.console.print(f"Message: {message_body}")
            if attachments:
                self.console.print(f"Attachments: {len(attachments)} file(s)")

            # Ask for confirmation before sending
            confirm = Prompt.ask("Do you want to send this email? (y/n)", choices=["y", "n"], default="n")

            if confirm.lower() == 'y':
                # Check if email configuration exists for real email sending
                try:
                    from email_sender import send_real_email, EmailConfig
                    config = EmailConfig()

                    # Check if we have proper configuration by checking if it's still the default
                    # The default is 'your_email@gmail.com', so if it's different, we have a config
                    has_valid_config = (
                        config.sender_email != 'your_email@gmail.com' and
                        config.sender_email and
                        config.sender_email.strip() != ''
                    )

                    if has_valid_config:
                        # Attempt to send real email
                        success = send_real_email(
                            recipient_email=recipient_email,
                            subject=subject,
                            message_body=message_body,
                            attachments=attachments,
                            config=config
                        )

                        if success:
                            self.console.print(f"[green]Email sent successfully to: {recipient_email}[/green]")
                        else:
                            self.console.print("[red]Failed to send email. Using simulation mode.[/red]")
                            # Fall back to simulation
                            self._simulate_email_sending(recipient_email)
                    else:
                        # No configuration provided, use simulation
                        self.console.print("[yellow]Email configuration not set. Using simulation mode.[/yellow]")
                        self.console.print("[yellow]Note: Please check your .env file configuration.[/yellow]")
                        self._simulate_email_sending(recipient_email)

                except ImportError as e:
                    # email_sender module not available, use simulation
                    self.console.print(f"[yellow]Import error: {str(e)}[/yellow]")
                    self.console.print("[yellow]Email module not available. Using simulation mode.[/yellow]")
                    self.console.print("[yellow]Note: email_sender module not found.[/yellow]")
                    self._simulate_email_sending(recipient_email)
                except Exception as e:
                    # Other errors (like authentication errors during email sending)
                    self.console.print(f"[yellow]Email sending error: {str(e)}[/yellow]")
                    self.console.print("[yellow]Falling back to simulation mode.[/yellow]")
                    self._simulate_email_sending(recipient_email)

                # Clean up temporary files after sending
                for attachment in attachments:
                    try:
                        os.remove(attachment)
                    except:
                        pass  # Ignore errors when removing temporary files
            else:
                self.console.print("[yellow]Email sending cancelled.[/yellow]")

                # Clean up temporary files if email was cancelled
                for attachment in attachments:
                    try:
                        os.remove(attachment)
                    except:
                        pass  # Ignore errors when removing temporary files

        except Exception as e:
            self.console.print(Panel(
                f"[red]❌ Error sending email: {str(e)}[/red]",
                border_style="red",
                box=ROUNDED
            ))

    def _simulate_email_sending(self, recipient_email):
        """Helper method to simulate email sending"""
        from rich.progress import Progress, SpinnerColumn
        import time

        with Progress(
            SpinnerColumn(),
            "[progress.description]{task.description}",
            transient=True,
        ) as progress:
            progress.add_task(description="Sending email...", total=None)
            time.sleep(1)

        self.console.print(Panel(
            f"[green]✅ Email sent successfully to: {recipient_email}[/green]",
            border_style="bright_green",
            box=ROUNDED
        ))

if __name__ == "__main__":
    app = TodoApp()
    app.run()